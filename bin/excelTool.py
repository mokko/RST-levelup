'''
Proof of concept version for ExcelTool

(1) create an updatable index in Excel

For our purposes, an index is 
-a list of distinct values
-typically in alphabetical order (but not necessarily)
-plus the number of occurrences of the values
-the data typically comes from an individual field, but I'd like to allow any xpath expression which includes
     ./sammlungsobjekt/maßangaben[@typ ='Ausgabe']

What do I mean with "updatable"?
The index is written to an excel file, user can edit this file, next run is supposed to update (and not destroy/overwrite) the index.

Later configuration comes from config file. At the moment we write the config into the class 

At the moment I use mpx for input, later I want to use LIDO for input .

    tool=ExcelTool (source_fn)
    tool.index ('mpx:museumPlusExport/mpx:sammlungsobjekt/mpx:sachbegriff)
        creates ./index.xslx with sheet "sachbegriff" which has index.


Excel Format
    Row 1: headers
    Column A: Gewimmel/Begriffe/Ausdrücke 
    Column B: Occurences

TODO: We assume that terms in excel are unique. I should check if that is the case.

'''

import os
from os import path
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook

#from cx_Freeze.samples.openpyxl.test_openpyxl import wb

class ExcelTool: 
    def __init__ (self, source, xls_fn):
        self.ns = {
            'npx': 'http://www.mpx.org/npx', #npx is no mpx
            'mpx': 'http://www.mpx.org/mpx', 
        }
        self.new_file=0
        self.tree = ET.parse(source)
        self.wb=self._prepare_wb(xls_fn)
        self.xls_fn=xls_fn

    #TODO
    def _xpath2core (self,xpath):
        '''this transformation is insufficient for a lot of expressions eg. involving [@attribute], but it'll do for the moment.
        '''
        core=xpath.split('/')[-1]
        core=core.split(':')[1].replace('[','').replace(']','').replace(' ','').replace('=','').replace('\'','')
        print ('xpath->core: ' + xpath + '->' + core)
        return core     
    
    
    def _prepare_ws (self, wb, xpath): 
        core=self._xpath2core(xpath) #extracts keyword from xpath for use as sheet.title

        try:
            ws = wb[core]
            return ws  #Sheet exists already, just return it

        except: 
            if self. new_file == 1:
                ws=wb.active
                ws.title=core
                self.new_file = None
                return ws
            else:
                return wb.create_sheet(core)

 
    def _prepare_header (self, ws):
        '''If Header columns are empty, fill them with default values'''
        if ws['A1'].value is None:
            ws['A1']='GEWIMMEL'
        if ws['B1'].value is None:
            ws['B1']='HÄUFIGKEIT'
        if ws['C1'].value is None:
            ws['C1']='QUALI'
        if ws['D1'].value is None:
            ws['D1']='PREF (DE)'
        if ws['E1'].value is None:
            ws['E1']='PREF (EN)'

    def _prepare_wb (self, xls_fn):
        '''Read existing xls or make new one, return values in self'''

        if path.isfile (xls_fn):
            print ('File exists, read it ('+ xls_fn+')')
            return load_workbook(filename = xls_fn)
        else:
            print ('File doesn\'t exist yet, making it ('+ xls_fn+')')
            self.new_file=1
            return Workbook()

    def _col_to_zero (self,ws,col):    
        '''
        Set all existing values of a respective column to 0. Only header (row=1) remains unchanged.
        
        self._col_to_zero (ws, 'B')
        '''
        c=1 # 1-based line counter 
        for each in ws[col]:
            if c != 1: #IGNORE HEADER
                #print (str(c)+': '+each.value)
                each.value=0 # None doesn't work
            c+=1
        return c


    def _del_col (self, ws, col):
        '''
        Delete all values in the respective column, column stays where it is. Also
        header (row=1) remains.
        
        self._del_col (ws, 'B')
        '''
        c=1 # 1-based line counter 
        for each in ws[col]:
            if c != 1: #IGNORE HEADER
                #print (str(c)+': '+each.value)
                each.value='' # None doesn't work
            c+=1
        return c

    #TODO -> broken identity test
    def index_with_attribute (self, xpath, quali): 
        '''Based on xpath determine sheet name and write vocabulary index to that xls sheet 
        
        TODO: I would like to allow a qualifier that describes the term
        Indien (Land) where "Land" is the qualifier for the term "Indien".
        
        But that's not so easy. I would need to prepare for the fact that the date could have both
        Indien (Land) and Indien (Subkontinent), i.e. same value, but different qualifiers. 
        Then the test for identity (_term_exist -> _term_exists_with_qualifier) needs to be different
        
        Maybe a cheap first version would not do that.
        '''
        ws=self._prepare_ws(self.wb, xpath)
        #print ('ws.title: '+ws.title)
        self._prepare_header(ws)
        self._col_to_zero(ws, 'B')

        for term in self.tree.findall(xpath, self.ns): 
            row=self._term_exists(ws, term.text)
            if row: 
                #print ('term exists already: '+str(row))
                cell='B'+str(row)
                value=ws[cell].value
                if value=='':
                    ws[cell]=1
                else:
                    ws[cell]=value+1
            else:
                print ('new term: '+ term.text)
            
            qu=term.get(quali)
            print ('QUALI: '+ quali+': '+ str(qu))
            self.insert_alphabetically(ws, term.text, qu)
                

        self.wb.save(self.xls_fn) 

    def index (self, xpath):
        '''Based on xpath determine sheet name and write vocabulary index to that xls sheet 
        '''

        ws=self._prepare_ws(self.wb, xpath)
        print ('ws.title: '+ws.title)
        self._prepare_header(ws)
        self._col_to_zero(ws, 'B') #drop col B with occurrences every time we run a new index

        for term in self.tree.findall(xpath, self.ns): 
            row=self._term_exists(ws, term.text)
            if row: 
                #print ('term exists already: '+str(row))
                cell='B'+str(row)
                value=ws[cell].value
                if value=='':
                    ws[cell]=1
                else:
                    ws[cell]=value+1
            else:
                print ('new term: '+ term.text)
                #append is only temporary, we want to sort it alphabetically for the most part #ws.append([term.text])  
                #ws.append({'A' : term.text})
                self.insert_alphabetically(ws, term.text)
                #print (term.text+': '+str(self.insert_alphabetically (ws, term.text)))

        self.wb.save(self.xls_fn) 


    def _line_alphabetically (self, ws, needle_term):
        '''CAVEAT: Uppercase and lowercase in alphabetial order ignored'''
        c=1 # 1-based line counter 
        for xlsterm in ws['A']:
            if c != 1: #IGNORE HEADER
                #print (str(c)+': '+each.value)
                if  needle_term.lower() < xlsterm.value.lower():
                    return c #found
                    #print (each.value + ' is before ' + needle_term)
                #else:
                    #print (each.value + ' is NOT before ' + needle_term)
            c+=1
        return c


    def insert_alphabetically (self, ws, term, quali=None): 
        '''
        inserts term into column A of worksheet ws after the first existing term
        
        ex: if we have list A,B,C, we want to put Ba between B und C
        
        looping current terms from xls
        each time comparing new term vs xls term
        needle_term is after first term
        needle_term is after second term
        needle_term is BEFORE third term -> so return a 2
        '''
        line=self._line_alphabetically(ws, term)
        ws.insert_rows(line)
        ws['A'+str(line)]=term
        ws['B'+str(line)]=1
        ws['C'+str(line)]=quali
        
        #print ('...insert at line '+str(line))


    def _term_exists (self, ws, needle_term):
        '''tests whether term is already in wb column A1
        ignores first row assuming it's a header and returns row of first occurrence
        
        row=self._term_exists(ws, 'Digitale Aufnahme')
        
        if row:
            print ('term exists already')
        else:
            print ('term new')
            
        if self._term_exists(ws, 'Digitale Aufnahme'):
            print ('term exists already')
        else:
            print ('term is new')
            
        
        '''
        c=1 # 1-based line counter 
        for each in ws['A']:
            if c != 1: #IGNORE HEADER
                #print (str(c)+': '+each.value)
                if each.value==needle_term:
                    return c #found
            c+=1
        return 0 #not found
        
        #wb.save(xls_fn)

if __name__ == '__main__': 
    t=ExcelTool ('data/WAF55/20190927/2-MPX/levelup.mpx', 'index.xlsx')
    t.index("./mpx:sammlungsobjekt/mpx:sachbegriff")
    t.index_with_attribute("./mpx:sammlungsobjekt/mpx:geogrBezug","bezeichnung")
    t.index("./mpx:sammlungsobjekt/mpx:geogrBezug[@bezeichnung = 'Ethnie']")
    t.index("./mpx:sammlungsobjekt/mpx:geogrBezug[@bezeichnung = 'Land']")
    