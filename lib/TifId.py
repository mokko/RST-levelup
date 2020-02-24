''' 
(1) Examine tifs on hard disk
(2) compare with mpx
(3) write info in xls for manual proof reading and as persistent storage
'''

import os, lxml
from openpyxl import Workbook, load_workbook
from pathlib import Path
from lxml import etree


class TifId: 
    def __init__(self, lib_fn):
        if not lib_fn.endswith('.xlsx'):
            raise ValueError ('Supplied Excel file name does not end on xlsx!')
        self.lib_fn=lib_fn
        self._prepare_wb(lib_fn)


    def scan_tif (self, tif_dir):
        '''Scan (recursively) for tif files in tif_dir and write results to excel library'''
        print ('* About to scan %s' % tif_dir)
        ws = self.wb.worksheets[0]
        #print (ws.title)

        for path in Path(tif_dir).rglob('*tif*'):
            if os.path.isfile (path): #we want only files, no dirs
                self._add_to_col(0, 'A', os.path.abspath(path))
                print(path)
        self._save_xsl()


    def mpx (self, mpx_fn):
        self._mpx(mpx_fn)
        self._save_xsl()


    def process_path (self):
        self._extract_identNr()
        self._xmp()
        self._save_xsl()


    '''PRIVATE METHODS'''
    '''Given a sheet with full path in col A, go thru that column and write the base filename in col B'''

    def _add_to_col (self, sheet_no, column, value):
        '''add value at first empty cell in given column'''
        ws = self.wb.worksheets[sheet_no]
        new_row=ws.max_row+1
        cell='%s%i' % (column, new_row)
        ws[cell]=str(value)


    def _extract_EM (self, path):
        base=os.path.basename(path)
        trunk,ext=os.path.splitext(base)
        ls=trunk.split()[:3]
        new=(' ').join(ls) # but actually has 4 parts TODO
        #print ('%s->%s' % (path, new))
        return new


    def _extract_identNr(self):
        print ('* Extract base')
        ws = self.wb.worksheets[0] #zero based!
        #print ('sheet title: %s' % ws.title)

        if ws.max_row > 1: # not with empty xlsx
            for col in ws.iter_cols(min_row=2, max_col=1, max_row=ws.max_row):
                for cell in col:
                    identNr=self._extract_EM(cell.value)
                    #print ('%i:%s' % (cell.row, identNr))
                    new_cell='B%i' % cell.row
                    if ws[new_cell].value is None:
                        #print ('Empty cell')
                        ws[new_cell]=identNr


    def _lookup (self, ET, needle):
        xpath="/m:museumPlusExport/m:sammlungsobjekt[m:identNr = '%s']/@objId" % needle
        print ('for needle %s' % needle)
        #print (xpath)
        #'/m:museumPlusExport/m:sammlungsobjekt[m:identNr = \'V A 695\']/@objId'
        r = ET.xpath(xpath, namespaces={'m':'http://www.mpx.org/mpx'})
        if len(r) == 1:
            print('-> %s' % r[0])
            return int(r[0])
        elif len(r) == 0:
            print ('nada')
            return False
        else:
            print (r)
            raise ValueError ('More than one objId todo')


    def _mpx (self, fn):
        '''Loop up objId by identNr in a mpx file'''
        print ('* Looking up objId in mpx')
        tree = etree.parse(fn)
        ws = self.wb.worksheets[0] #zero based!

        if ws.max_row > 1: # not with empty xlsx
            for col in ws.iter_cols(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
                for cell in col:
                    identNr=cell.value
                    objId=self._lookup(tree, identNr)
                    if objId is True and ws[cell].value is None:
                        cell='D%i' % cell.row
                        ws[cell]=objId


    def _prepare_wb (self, xls_fn):
        if os.path.isfile (xls_fn):
            #print ('File exists ('+ xls_fn+')')
            self.wb=load_workbook(filename = xls_fn)
        else:
            print ("Excel library file doesn't exist yet, making it ")
            #raise ("Excel file not found: %s" % xls_fn)
            self.wb=Workbook()
            ws1 = self.wb.active
            ws1.title = "tifs"
            ws1['A1'] = 'Pfad'
            ws1['B1'] = 'Signatur aus Pfad'
            ws1['C1'] = 'IdentNr aus MPX'
            ws1['D1'] = 'objId aus MPX'
            ws1['E1'] = 'Fotograf aus Tif'
            ws1['F1'] = '(C) aus Tif'
            #self.wb.create_sheet (title="jpgs") -> comparison with jpg is no longer necessary


    def _save_xsl (self):
        print ('* Saving excel library to %s' % self.lib_fn)
        try:
            self.wb.save(self.lib_fn)
        except:
            print ('Saving exception caught')

    def _xmp (self):
        import exifread
        print ('* Extract from tiff (xmp)')
        ws = self.wb.worksheets[0] #zero based!
    
        if ws.max_row > 1: # not with empty xlsx
            for col in ws.iter_cols(min_row=2, max_col=1, max_row=ws.max_row):
                for cell in col:
                    path = os.path.realpath(cell.value)
                    print (path)
                    if os.path.isfile (path):
                        print (path)
                        E_cell='E%i' % cell.row
                        F_cell='F%i' % cell.row
                        if ws[E_cell].value is None or ws[F_cell].value is None: 
                            f = open(path, 'rb')
                            tags = exifread.process_file(f)
                            if 'Image Artist' in tags:
                                ws[E_cell]=str(tags['Image Artist'])
                                print ('    %s' % str(tags['Image Artist']))
                            if 'Image Copyright' in tags:
                                ws[F_cell]=str(tags['Image Copyright'])
                                print ('    %s' % str(tags['Image Copyright']))
                        else:
                            print ('Debug: No file %s' % path)
