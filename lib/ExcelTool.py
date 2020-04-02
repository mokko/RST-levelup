"""Creates and updates vocabulary indexes in an Excel file.

Can also replaces syns with prefs.

(1) Expects configuration file (json)
(2) Extracts terms from xml source
(3) it creates/updates an XLS file with the vocabulary index
(4) creates another xls for translation purposes
(5) can replace syns with prefs into a new xml file 

For our purposes, a vocabulary index is an alphabetical list of terms with 
their frequency in the source data. The terms are regarded as synonyms and 
associated with preferred terms. Preferred terms can be replace synonyms
in the data as a way of cleaning up the data.

Excel Format
    Row 1: headers
    Column A: Gewimmel/Begriffe/Ausdrücke
    Column B: Qualifier
    Column C: Division
    Column C: Frequency

TODO: We assume that terms in excel are unique. They are unique when we first
write them, but user could create non-unique terms. I could check if 
uniqueness.

This class is agnostic as to where you locate conf_fn, but I recommend
    data/EM-SM/vindex-config.json
    data/EM-SM/20200113/2-MPX/levelup.mpx

USAGE
    #low level constructor
    t=ExcelTool (source_xml, [xls_dir])
    t.index ('mpx:museumPlusExport/mpx:sammlungsobjekt/mpx:sachbegriff)
        creates or updates ./index.xlsx with sheet "sachbegriff" 

    #high level constructor (using commands from conf file)
    t=ExcelTool.from_conf (conf_fn,source_fn) # runs the commands in the conf_fn
    t.apply_fix (conf_fn, out_fn) # writes cleanup version to out_fn
TODO: new translation feature (
: are triggered in from_conf
: instructions are in conf.json
: target_dir specified during from_conf"""

import os
from lxml import etree #has getParent()
from openpyxl import Workbook, load_workbook


class ExcelTool:

    def __init__ (self, source_xml,xls_dir = '.'):
        self.ns = {
            'npx': 'http://www.mpx.org/npx', #npx is no mpx
            'mpx': 'http://www.mpx.org/mpx', 
        }
        self.tree = etree.parse(source_xml)
        self.new_file = 0
        self.xls_fn = os.path.relpath(os.path.realpath(os.path.join (xls_dir,'vindex.xlsx')))
        self.wb = self._prepare_wb(self.xls_fn) 
        self.trans_xls = os.path.relpath(os.path.realpath(os.path.join (xls_dir,'translate.xlsx')))
        self.twb = self._prepare_wb(self.trans_xls)

    def apply_fix (self, conf_fn, out_fn):
        """Replace syns with prefs in out_fn
        
        Preparation: read three files
        1. Read conf and process every task
        2. Read xls_fn and use matching sheet
        3. Read xml/mpx source 
        Rewrite 
        4. mpx according to instructions from conf and prefs in xls
        5. save new xml/mpx to out_fn"""

        print ('*About to apply vocabulary control')
        
        #primitive Domain Specific Language (DSL)
        for task, cmd in self._itertasks(conf_fn):
            xpath = task[cmd][0]
            if cmd == 'index': pass
            elif cmd == 'index_with_attribute':
                attribute = task[cmd][1]
            elif cmd == 'index_with_2attributes': pass
            elif cmd == 'attribute_index':
                include_verant = task[cmd][1]
                xpath, attrib = self._attribute_split(task[cmd][0]) #rewrite xpath for iterterms
            elif cmd == "translate_element": pass
            elif cmd == "translate_attribute": pass
            else:
                print (f"WARNING: Unknown command found in conf {cmd}")
            ws = self._get_ws (task[cmd][0]) 

            if (cmd == 'index' 
                or cmd == 'index_with_attribute'
                or cmd == 'index_with_2attributes'
                or cmd == 'attribute_index'):
                print (f"**Checking replacements from sheet '{ws.title}'")
                print (f"   {cmd}: {task[cmd]}")
    
                for term, verant in self._iterterms(xpath):
                    term_str=self._term2str (term) #strip whitespace
                    if cmd == 'index': 
                        lno = self._term_verant_exists(ws, term_str, verant)
                        #print(f"syn term found '{term.text}' {lno}")
                    elif cmd == 'index_with_attribute':
                        qu_value = self._get_attribute (term, attribute) 
                        lno = self._term_quali_exists(ws, term_str, qu_value, verant)
                        #print(f"syn term found '{term.text}' {lno}")
                    elif cmd == 'index_with_2attributes':
                        quv1 = self._get_attribute (term, task[cmd][1]) 
                        quv2 = self._get_attribute (term, task[cmd][2]) 
                        qu_value=f"{quv1} - {quv2}"
                        lno = self._term_quali_exists(ws, term_str, qu_value, verant)
                    elif cmd == 'attribute_index':
                        try:
                            value=term.attrib[attrib]
                        except: pass
                        if include_verant == 'verantwortlich':
                            lno = self._term_verant_exists(ws, value, verant)
                        else:
                            lno = self._term_exists(ws, value)
                        #print(f"syn attribute found {value} {lno}")

                    if lno: # no replace if term is not in xls
                        pref_de = ws[f'E{lno}'].value
                        if pref_de is not None: #no replace if pref is not given
                            #print (f"pref found: {pref_de}")
                            if cmd == 'index' or cmd == 'index_with_attribute': #if value?
                                print (f"   replace term: {term_str}->{pref_de}")
                                term.text = pref_de.strip() # modify xml
                            else:
                                print (f"   replace attribute '{attrib}': {value}->{pref_de}")
                                term.attrib[attrib] = pref_de.strip() # modify xml

        print (f"*About to write xml to {out_fn}")
        #register_namespace('', 'http://www.mpx.org/mpx') #why? default ns?
        self.tree.write(out_fn, encoding="UTF-8", xml_declaration=True)

    def translate_from_conf (conf_fn, source_xml, xls_dir = None):
        """constructor or regular method?
        
        New version as constructor to be more expectable."""

        if xls_dir is None:
            xls_dir=os.path.dirname(conf_fn)
        t=ExcelTool (source_xml,xls_dir)

        for task,cmd in t._itertasks(conf_fn): #sort of a Domain Specific Language DSL
            if cmd == "translate_element": 
                t.translate_element (task[cmd])
            elif cmd == "translate_attribute": 
                t.translate_attribute (task[cmd])
            elif cmd == "index_with_attribute": pass
            elif cmd == "attribute_index": pass
            elif cmd == "index": pass
            elif cmd == "index_with_2attributes": pass
            else: 
                print (f"WARNING: Unknown command in conf {cmd}")
        return t

    def from_conf (conf_fn, source_xml, xls_dir = None): #no self
        """Constructor that executes commands from conf_fn"""
        # as default use the same dir for xls as for conf_fn
        if xls_dir is None:
            xls_dir=os.path.dirname(conf_fn)
        t=ExcelTool (source_xml,xls_dir)

        for task,cmd in t._itertasks(conf_fn): #sort of a Domain Specific Language DSL
            #print (f"from_conf: {cmd}: {task[cmd]}")
            if cmd == "index":
                t.index (task[cmd][0], task[cmd][1])
            elif cmd == "index_with_attribute":
                t.index_with_attribute (task[cmd][0], task[cmd][1])
            elif cmd == "index_with_2attributes":
                t.index_with_2attributes (task[cmd][0], task[cmd][1], task[cmd][2])
            elif cmd == "attribute_index":
                t.index_for_attribute (task[cmd][0], task[cmd][1])
            elif cmd == "translate_element": pass
            elif cmd == "translate_attribute": pass
            else:
                print (f"WARNING: Unknown command in conf {cmd}")
        return t

    def index (self, xpath, include_verant=''):
        """Write vocabulary index to the right xls sheet.

        Sheet depends on xpath expression."""

        print(f"**Creating/updating voc-index for element {xpath}")
        ws = self._prepare_indexing(xpath, self.wb)

        for term, verant in self._iterterms(xpath):
            term_str = self._term2str (term) #if there is whitespace we don't want it 
            if include_verant == 'verantwortlich':
                row = self._term_verant_exists(ws, term_str, verant)
            else:
                #print ("verantwortlich is NOT part of the identity test")
                verant=None
                row = self._term_exists(ws, term_str)
            if row: 
                #print ('term exists already: '+str(row))
                self._update_frequency (ws, row)
            else:
                print (f"new term: {term_str}")
                self._insert_alphabetically(ws, term_str, verant)
        self.wb.save(self.xls_fn) 

    def index_for_attribute (self, xpath, include_verant=''):
        """Make vocabulary index for an attribute
        
        Assuming the xpath expression ends with something like:
            mpx:bla/@attribute
        
        Once I have the attribute value I dont get back to parent. Even in 
        lxml."""

        print(f"**Creating/updating voc-index for attribute {xpath}")
        ws = self._prepare_indexing(xpath, self.wb)
        base_xpath, attrib = self._attribute_split(xpath)

        for term, verant in self._iterterms(base_xpath):
            value = term.get(attrib)
            if value is not None:
                #print (f"***Value {value}")
                if include_verant == 'verantwortlich':
                    row = self._term_verant_exists(ws, value, verant)
                else:
                    #print ("verantwortlich is NOT part of the identity test")
                    verant=None
                    row = self._term_exists(ws, value)
                if row:
                    self._update_frequency (ws, row)
                else:
                    print (f"new attribute: {value}")
                    self._insert_alphabetically(ws, value, verant)
        self.wb.save(self.xls_fn) 

    def index_with_2attributes (self, xpath, quali1, quali2): 
        """Write vocabulary index for an element with 2 qualifiers

        Treats terms with different qualifiers as two different terms, e.g. 
        lists both Indien (Land) and Indien ()."""

        print(f"**Creating/updating voc-index for element with 2attributes {xpath} {quali1} {quali2}")
        ws = self._prepare_indexing(xpath, self.wb)

        for term, verant in self._iterterms(xpath):
            qu_value1 = self._get_attribute(term, quali1)
            qu_value2 = self._get_attribute(term, quali2)
            #print (f"{quali1}:{qu_value1}; {quali2}:{qu_value2}")
            term_str = self._term2str (term) #no whitespace 
            qu_value=f"{qu_value1} - {qu_value2}"
            #print (f"{qu_value}")
            row = self._term_quali_exists(ws, term_str,qu_value, verant)
            if row:
                #print ('term exists already: '+str(row))
                self._update_frequency (ws, row)
            else:
                print (f'new term: {term_str} ({qu_value})')
                self._insert_alphabetically(ws, term_str, verant, qu_value)
        self.wb.save(self.xls_fn) 
        
    def index_with_attribute (self, xpath, quali): 
        """Write vocabulary index for an element with qualifier
        Treats terms with different qualifiers as two different terms, e.g. 
        lists both Indien (Land) and Indien ()."""

        print(f"**Creating/updating voc-index for element with attribute {xpath}")
        ws = self._prepare_indexing(xpath, self.wb)

        for term, verant in self._iterterms(xpath):
            qu_value = self._get_attribute(term, quali)
            term_str = self._term2str (term) #no whitespace 
            row = self._term_quali_exists(ws, term_str,qu_value, verant)
            if row:
                #print ('term exists already: '+str(row))
                self._update_frequency (ws, row)
            else:
                print (f'new term: {term_str} ({qu_value})')
                self._insert_alphabetically(ws, term_str, verant, qu_value)
        self.wb.save(self.xls_fn) 
        
    def translate_attribute (self, xpath):
        """Write/update translation xls for attribute"""

        print(f"*Creating/updating translation sheet for attribute {xpath}")
        ws = self._prepare_ws(xpath, self.twb)
        print (f"   sheet {ws.title}")
        self._prepare_header_trans(ws)
        self._col_to_zero(ws, 'D') #drop all frequencies and begin again
        base_xpath, attrib = self._attribute_split(xpath)
        for term, verant in self._iterterms(base_xpath): 
            value = term.get(attrib)
            if value is not None:
                row = self._term_exists (ws, value)
                if row:
                    self._update_frequency (ws, row)
                else:
                    print (f"new term: {term.text}")
                    self._insert_alphabetically(ws, value)
        self.twb.save(self.trans_xls) 

    def translate_element (self, xpath):
        """Write/update translation xls based on source_xml"""

        print(f"*Creating/updating translation sheet for {xpath}")
        ws = self._prepare_ws(xpath, self.twb)
        self._prepare_header_trans(ws)
        self._col_to_zero(ws, 'D') #drop all frequencies and begin again
        print (f"   sheet {ws.title}")
        for term, verant in self._iterterms(xpath): 
            row = self._term_exists (ws, term.text)
            if row:
                self._update_frequency (ws, row)
            else:
                print (f"new term: {term.text}")
                self._insert_alphabetically(ws, term.text)
        self.twb.save(self.trans_xls) 

        
#    PRIVATE STUFF

    def _attribute_split (self, xpath):
        attrib = xpath.split('/')[-1]
        if attrib.startswith("@"):
            elems = xpath.split('/')[:-1]
            main_xpath = '/'.join(elems)
        else:
            raise ValueError(f"Error: Expect attribute in last position: {xpath}")
        return main_xpath, attrib[1:]

    def _col_to_zero (self,ws,col):
        """Set all values of a specific column to 0. 
        
        Only header (row=1) remains unchanged.
        
        USAGE: 
            self._col_to_zero (ws, 'B')"""

        c=1 # 1-based line counter 
        for each in ws[col]:
            if c != 1: #IGNORE HEADER
                #print (str(c)+': '+each.value)
                each.value=0 # None doesn't work
            c+=1
        return c

    def _del_col (self, ws, col):
        """ Delete all values in a specific column. 
        
        Header (row=1) remains as is.
        
        USAGE:
            self._del_col (ws, 'B')"""

        c=1 # 1-based line counter 
        for each in ws[col]:
            if c != 1: #IGNORE HEADER
                #print (str(c)+': '+each.value)
                each.value='' # None doesn't work
            c+=1
        return c

    def _get_attribute (self, node, attribute):
        value = node.get(attribute)
        if value is not None:
            return value.strip() #strip probably unnecessary when M+

    def _get_ws (self,xpath):
        """Get existing worksheet based on xpath or die
        
        Compare with _prepare_ws which doesn't die"""

        core=self._xpath2core(xpath) #extracts keyword from xpath for use as sheet.title
        return self.wb[core] # dies if sheet with title=core doesn't exist

    def _insert_alphabetically (self, ws, term, verant=None, quali=None): 
        """Inserts new term into column A alphabetically."""

        line=self._line_alphabetically(ws, term)
        ws.insert_rows(line)
        ws[f'A{line}'] = term
        ws[f'B{line}'] = quali
        ws[f'C{line}'] = verant
        ws[f'D{line}'] = 1 #this is a new term

    def _itertasks(self, conf_fn):
        data = self._read_conf(conf_fn)
        for task in data['tasks']:
            for cmd in task:
                yield task,cmd

    def _iterterms (self, xpath):
        """Finds all xpaths nodes and who is verantwortlich. 
        
        Assumes that verantwortlich is a sibling node."""

        for term in self.tree.findall(xpath, self.ns):
            verant_node = term.find("../mpx:verantwortlich", self.ns) #assuming that it always exists 
            try: 
                verant = verant_node.text
            except:
                verant = None
                #Im MM Modul gibt es keine Verantwortlichkeit
                #print ("*****niemand verantwortlich")

            if term is not None:
                yield term, verant

    def _line_alphabetically (self, ws, needle_term):
        """Assuming alphabetical sort, return line where term fits
        
        Uppercase and lowercase in order ignored."""

        lno=1 # 1-based line counter 
        for xlsterm in ws['A']:
            if lno > 1: #IGNORE HEADER
                if  needle_term.lower() < xlsterm.value.lower():
                    return lno #found
            lno += 1
        return lno #if needle not found, return 1

    def _prepare_indexing(self, xpath, wb):
        ws = self._prepare_ws(xpath, wb)
        self._prepare_header(ws)
        self._col_to_zero(ws, 'D') #drop all frequencies when working on index
        print (f"   sheet {ws.title}")
        return ws

    def _prepare_ws (self, xpath, wb):
        """Get existing sheet or make new one. 
        
        Sheet title is based on xpath expression."""

        sheet_label = self._xpath2core(xpath) 

        try:
            ws = wb[sheet_label]
        except: 
            if self.new_file == 1:
                ws = wb.active
                ws.title = sheet_label
                self.new_file = None
                return ws
            else:
                return wb.create_sheet(sheet_label)
        else:
            return ws #Sheet exists already, just return it

    def _prepare_header (self, ws):
        """Fill header columns with default values, if they are empty."""

        columns = {
            'A1': 'GEWIMMEL*',
            'B1': 'QUALI*', #create this column even if not used
            'C1': 'VERANTWORTLICHKEIT*',
            'D1': 'FREQUENZ*',
            'E1': 'VORZUGSBEZEICHNUNG',
            'F1': 'NOTIZEN'
        }
        self._write_header(columns,ws)

    def _prepare_header_trans (self, ws):
        columns = {
            'A1': 'DE*',
            'B1': 'EN',
            'C1': 'NOTIZEN',
            'D1': 'FREQUENZ*'
        }
        self._write_header(columns,ws)

    def _prepare_wb (self, xls_fn):
        """Read existing xls or make new one.
        
        Returns workbook."""

        if os.path.isfile (xls_fn):
            print (f'   Excel file exists ({xls_fn})')
            return load_workbook(filename = xls_fn)
        else:
            print (f"   Excel file doesn't exist yet, making it ({xls_fn})")
            self.new_file=1
            return Workbook()

    def _read_conf (self, conf_fn):
        import json
        with open(conf_fn, encoding='utf-8') as json_data_file:
            data = json.load(json_data_file)
        return data

    def _term2str (self, term_node):
        term_str = term_node.text
        if term_str is not None: #if there is whitespace we want to ignore it in the index
            return term_str.strip()

    def _term_exists (self, ws, term):
        """Tests if the term exists already.

        Ignores first row assuming it's a header. Returns row of first 
        occurrence."""

        lno=1 # 1-based line counter 
        for each in ws['A']:
            if lno > 1: #IGNORE HEADER
                if each.value == term:
                    #print(f"{each.value} ({verant}) == {term} ({ws[f'C{lno}'].value})")
                    return lno #found
            lno+=1
        return 0 #term not found

    def _term_verant_exists (self, ws, term, verant):
        """Tests if the combination of term and verantwortlich exists already.

        Should we include verantwortlich in identity check?

        Ignores first row assuming it's a header. Returns row of first 
        occurrence."""

        lno=1 # 1-based line counter 
        for each in ws['A']:
            if lno > 1: #IGNORE HEADER
                if each.value == term and ws[f'C{lno}'].value == verant:
                    #print(f"{each.value} ({verant}) == {term} ({ws[f'C{lno}'].value})")
                    return lno #found
            lno+=1
        return 0 #term not found

    def _term_quali_exists(self,ws, term,quali, verant):
        """Tests if the combination of term/qualifier/verantwortlich exists.

        Returns 0 if combination not found. Otherwise, returns line number 
        of first occurrence. 

        SEE ALSO: _term_exists

        If user deletes verantwortlich anywhere in Excel file, program will 
        die."""

        lno=1 # 1-based line counter 
        for each in ws['A']:
            if lno != 1: #IGNORE HEADER
                #print (f"{c}: {each.value}")
                if (each.value == term 
                    and ws[f'B{lno}'].value == quali 
                    and ws[f'C{lno}'].value == verant) :
                    return lno #found
            lno+=1
        return 0 #not found

    def _update_frequency (self, ws, row_no):
        """Adds one to frequency column"""

        cell = f"D{row_no}" #frequency in column D 
        value = ws[cell].value
        if value == '':
            ws[cell] = 1
        else:
            ws[cell] = value + 1

    def _write_header (self, columns, ws):
        from openpyxl.styles import Font
        for key in columns:
            if ws[key].value is None:
                ws[key] = columns[key]
                c = ws[key]
                c.font = Font(bold=True)

    def _xpath2core (self,xpath):
        """Take xpath and return a string suitable that works as a sheet title.
        
        This algorithm is pretty stupid, but it'll do for the moment."""

        core = xpath.split('/')[-1]
        if core.startswith('@'): #assumes that attributes don't have ns
            core = xpath.split('/')[-2] + core
        try:
            core = core.split(':')[1]
        except: pass
        core = core.replace('[','').replace(']','').replace(' ','').replace('=','').replace('\'','')
        if len(core) > 31:
            core=core[:24]+'...'
        #print (f"***xpath->core: {xpath} -> {core}")
        return core

if __name__ == '__main__': 
    #t=ExcelTool ('2-MPX/levelup.mpx', '.')
    #t.index("./mpx:sammlungsobjekt/mpx:personenKörperschaften")
    #t.index_for_attribute("./mpx:sammlungsobjekt/mpx:geogrBezug/@bezeichnung", "dont test and record verant")
    #paths expect that you run it from the date directory (e.g. 20200226)
    t=ExcelTool.from_conf('../vindex.json', '2-MPX/levelup.mpx')
    t.apply_fix('../vindex.json', 'test-fix.mpx')
