import os 
from openpyxl import Workbook, load_workbook

class XlsTools:
    def _prepare_wb (self, xls_fn):
        """Read existing xls or make new one.
        
        Returns workbook."""

        if os.path.isfile (xls_fn):
            #print (f'   Excel file exists ({xls_fn})')
            return load_workbook(filename = xls_fn)
        else:
            print (f"   Excel file doesn't exist yet, making it ({xls_fn})")
            return Workbook()
