""" Find *.tif? files by filename/identNr 

    Uses a json file as cache to store path information (e.g. .tif_finder.json)

    USAGE:
        tf=Tif_finder(cache_fn)
        tf.scandir(scan_dir) #scans recursively for *.tif?
        tf.show() #prints cache to STDOUT 

        ls=tf.search (needle) #matching path in list
        tf.search_xls(needle) #with needles from xls, search cache & report to
                              #STDOUT

        tf.search_xls(xls_fn, outdir) #search cache for needles from xls and 
                              #copy found tifs to outdir
        tf.search_mpx(mpx_fn, outdir) #search for all identNr in mpx and copy
                              #to outdir 

    
    Filename is usually preserved; only if multiple tifs have the same name
    they are varied by adding a number.

    The downside of this naming scheme is that if Tif_finder is run multiple 
    times, same files will be copied multiple times. For use inside a CLI util 
    that may still work, for other cases we will need different naming 
    convention.
    
    $objId.$hash.tif
    Before I program that, I want to see that I can actually find the right 
    objId RELIABLY."""

import shutil
import json
import os
import hashlib
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from lxml import etree


class Tif_finder:
    def __init__(self, cache_fn): 
        """initialize object

        cache_fn: location (path) of cache file"""
        self.cache_fn=cache_fn
        #print ('cache_fn %s' % cache_fn)

        if os.path.exists(self.cache_fn):
            print (f"*cache exists, loading '{self.cache_fn}'")
            with open(self.cache_fn, 'r') as f:
                self.cache = json.load(f)
        else:
            self.cache={}


    def scandir (self, scan_dir):
        """Scans directory for *.tif|*.tiff files recursively and writes
        results to cache file, updating existing cache file or starting new 
        one.

        scan_dir needs to be a directory.

        Does a sloppy update, i.e will not remove cache entries for files that
        have been removed from disk.

        Scan multiple directories by running the scan multiple times.
        
        TODO: I could scan all items in cache and check if they still exist
        on disk, and delete cache items accordingly. to make update not 
        sloppy."""

        if not os.path.isdir (scan_dir):
            raise ValueError (f"Scan dir '{scan_dir}' does not exist")

        print (f"* About to scan {scan_dir}")
        files=Path(scan_dir).rglob('*.tif') # returns generator
        files2=Path(scan_dir).rglob('*.tiff')
        for path in list(files) +list(files2):
            abs = path.resolve()
            base = os.path.basename(abs)
            (trunk,ext)=os.path.splitext(base)
            print (f"{abs}")
            self.cache[str(abs)]=str(trunk)

        print ('* Writing updated cache file')
        with open(self.cache_fn, 'w') as f:
            json.dump(self.cache, f)


    def search (self, needle, target_dir=None):
        """Search cache for a single needle, returns list of matches
            ls=self.search(needle)"""

        #print ("* Searching cache for needle '%s'" % needle)
        ret=[path for path in self.cache if needle in self.cache[path]]

        if target_dir is not None:
            for f in ret:
                self._simple_copy(f,target_dir)

        return ret


    def search_xls (self, xls_fn, target_dir=None):
        """Search tifs for needles in Excel file (first sheet, column A)

        If target_dir is not None, copy the file to respective directory.
        If target_dir is None just report matching paths to STDOUT"""

        print (f"* Searching cache for needles from Excel file {xls_fn}")

        self.wb=self._prepare_wb(xls_fn)
        #ws = self.wb.active # last active sheet
        ws = self.wb.worksheets[0]
        print (f"* Sheet title: {ws.title}")
        col = ws['A'] # zero or one based?
        for needle in col:
            #print ('Needle: %s' % needle.value)
            if needle != 'None':
                found=self.search(needle.value)
                print(f'found {found}')
                if target_dir is not None:
                    for f in found:
                        self._simple_copy(f,target_dir)


    def search_mpx (self, mpx_fn, target_dir=None):
        """For each identNr from mpx look for corresponding tifs in cache and 
        copy them to target_dir"""
        
        if target_dir is not None:
            target_dir=os.path.realpath(target_dir)
            if not os.path.isdir(target_dir):
                os.makedirs (target_dir)
        tree = etree.parse(mpx_fn)
        r = tree.xpath('/m:museumPlusExport/m:sammlungsobjekt/m:identNr', 
            namespaces={'m':'http://www.mpx.org/mpx'})

        for identNr_node in r:
            tifs=self.search (identNr_node.text)
            objId=tree.xpath(f"/m:museumPlusExport/m:sammlungsobjekt/@objId[../m:identNr = '{identNr_node.text}']", 
                namespaces={'m':'http://www.mpx.org/mpx'})[0]
            #print(f"{identNr_node.text}->{objId}")
            found=self.search(identNr_node.text)
            for f in found:
                print (f"{identNr_node.text}->{objId}->{f}")
                if target_dir is not None:
                    self._hash_copy(f, target_dir, objId)


    def show(self):
        """Prints contents of cache to STDOUT"""

        print ('*Displaying cache contents')
        if hasattr (self, 'cache'):
            for item in self.cache:
                print (f"  {item}")
            print (f"Number of tifs in cache: {len(self.cache)}")
        else:
            print (' Cache does not exist!')


#############

    def _init_log (self,outdir):
        #line buffer so everything gets written when it's written, so I can CTRL+C the program
        self._log=open(outdir+'/report.log', mode="a", buffering=1)

    def _write_log (self, msg):
        self._log.write(f"[{datetime.datetime.now()}] {msg}\n")
        print (msg)

    def _close_log (self):
        self._log.close()


    def _target_fn (self, fn):
        """Return filename that doesn't exist yet. 
        
        Check if target exists and if so, find & return new variant that does 
        not yet exist according to the following schema:
            path/to/base.ext
            path/to/base (1).ext
            path/to/base (2).ext
            ..."""

        new=fn
        i=1
        while os.path.exists (new):
            #print ('Target exists already')
            trunk,ext=os.path.splitext(fn)
            new=f"{trunk} ({i}).{ext}"
            i+=1
        print (f"[{i}] {new}")
        return new


    def _simple_copy(self, source, target_dir):
        """Copy source file to target dir, typically keeping original 
        filename. Only if there already is a file with that name, find a new
        name that doesn't exist yet. 
        
        Upside: we can have multiple tifs for one identNr. 
        Downside: new filenames don't necessarily match the old one."""


        if not os.path.isdir(target_dir):
            raise ValueError ("Error: Target is not directory!")
        #print ('cp %s -> %s' %(source, target_dir))
        self._init_log(target_dir)
        s_base = os.path.basename(source)
        target_fn=self._target_fn(os.path.join(target_dir, s_base)) # should be full path
        if not os.path.isfile(target_fn): #no overwrite
            self._write_log(f"{source} -> {target_fn}") 
            try: 
                shutil.copy2(source, target_fn) # copy2 preserves file info
            except:
                self._write_log(f"File not found: {source}")
        self._close_log()


    def _hash_copy (self, source, target_dir, objId):
        """ Copy *.tif to target_dir/$objId.$hash.tif"""

        if not os.path.isdir(target_dir):
            raise ValueError ("Error: Target is not directory!")
        self._init_log(target_dir)
        hash=self._file_hash(source)
        target_fn=os.path.join(target_dir, f"{objId}.{hash}.tif")
        if not os.path.isfile(target_fn): #no overwrite
            self._write_log(f"{source} -> {target_fn}") 
            try: 
                shutil.copy2(source, target_fn) # copy2 preserves file info
            except:
                self._write_log(f"File not found: {source}")
        self._close_log()


    def _file_hash (self, fn):
        print (f"About to hash '{fn}'...", end='')
        with open(fn, "rb") as f:
            file_hash = hashlib.md5()
            while chunk := f.read(8192): #walrus operator requires python 3.8
                file_hash.update(chunk)
        print('done')
        return file_hash.hexdigest()


    def _prepare_wb(self, xls_fn):
        """Read existing xlsx and return workbook"""

        if os.path.isfile(xls_fn):
            #print (f"File exists ({xls_fn})")
            return load_workbook(filename = xls_fn)
        else:
            raise ValueError (f"Excel file not found: {xls_fn}")


if __name__ == "__main__":
    
    f=Tif_Finder()
